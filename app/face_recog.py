import cv2
import os
import numpy as np
from datetime import datetime
import openpyxl
import time
import sys

# Load Haar cascade
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Base directory
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# ------------------ TRAINING ------------------
def train_model(branch, section):
    faces = []
    labels = []
    label_map = {}
    label_id = 0

    folder_path = os.path.join(BASE_DIR, "register_candidates", f"{branch}_{section}")
    if not os.path.exists(folder_path):
        print(f"❌ Folder not found: {folder_path}")
        return None

    for dir_name in os.listdir(folder_path):
        person_path = os.path.join(folder_path, dir_name)
        if not os.path.isdir(person_path):
            continue

        print(f"📂 Loading from: {person_path}")
        label_map[label_id] = dir_name

        for img_file in os.listdir(person_path):
            img_path = os.path.join(person_path, img_file)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if img is not None:
                faces.append(img)
                labels.append(label_id)
            else:
                print(f"⚠️ Could not load image: {img_path}")

        label_id += 1

    if not faces:
        print("❌ No training data found!")
        return None

    recognizer.train(faces, np.array(labels))
    print(f"✅ Model trained with {len(faces)} images.")
    return label_map

# ------------------ ATTENDANCE ------------------
def mark_attendance(record_name, subject, branch, section, status):
    roll, name = record_name.split('_', 1)
    date_str = datetime.now().strftime('%Y-%m-%d')
    time_str = datetime.now().strftime('%H:%M:%S')

    attendance_dir = os.path.join(BASE_DIR, "attendance", date_str)
    os.makedirs(attendance_dir, exist_ok=True)

    file_path = os.path.join(attendance_dir, f"{subject}_{branch}_{section}.xlsx")

    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Roll No", "Name", "Branch", "Section", "Date", "Time", "Status"])
        wb.save(file_path)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(values_only=True):
        if row[0] == roll and row[4] == date_str:
            print(f"🟡 Attendance already marked for {roll}")
            return

    sheet.append([roll, name, branch, section, date_str, time_str, status])
    wb.save(file_path)
    print(f"✅ Attendance marked in {file_path}")

# ------------------ RECOGNITION ------------------
def recognize_faces(label_map, subject, branch, section):
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    if not cap.isOpened():
        print("❌ Webcam not accessible!")
        return

    recognized = False
    start_time = time.time()

    while True:
        ret, frame = cap.read()
        if not ret:
            print("⚠️ Failed to capture frame.")
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.2, 5)

        for (x, y, w, h) in faces:
            face_img = gray[y:y+h, x:x+w]
            face_img = cv2.resize(face_img, (200, 200))

            try:
                label, confidence = recognizer.predict(face_img)
                if confidence < 60:
                    person = label_map[label]
                    print(f"✅ {person} recognized with confidence {confidence:.2f}")
                    mark_attendance(person, subject, branch, section, "Present")

                    cv2.putText(frame, f"{person} - Present", (x, y - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                    recognized = True
                    break
                else:
                    print("❌ Unrecognized face or low confidence")
                    cv2.putText(frame, "Invalid", (x, y - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
            except Exception as e:
                print(f"❌ Error during recognition: {e}")

        cv2.imshow("Recognizing... Press 'q' to quit", frame)

        if recognized or cv2.waitKey(1) & 0xFF == ord('q'):
            break

        if time.time() - start_time > 15:
            print("⏰ Timeout: No face recognized within time limit.")
            break

    cap.release()
    cv2.destroyAllWindows()

# ------------------ MAIN ENTRY ------------------
if __name__ == '__main__':
    try:
        if len(sys.argv) == 4:
            subject = sys.argv[1].strip()
            branch = sys.argv[2].strip().lower()
            section = sys.argv[3].strip().lower()
        else:
            print("❌ Error: Please provide subject, branch, and section as arguments.")
            print("✅ Example: python face_recog.py Math cse b1")
            sys.exit(1)

        label_map = train_model(branch, section)
        if label_map:
            recognize_faces(label_map, subject, branch, section)
        else:
            print("❌ Could not proceed due to missing training data.")
    except Exception as e:
        print(f"❌ Error: {e}")
