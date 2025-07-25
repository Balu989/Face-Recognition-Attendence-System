import cv2
import os
import sys
import numpy as np
from datetime import datetime
import openpyxl

face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()


def train_model(folder_path):
    faces, labels = [], []
    label_map = {}
    label_id = 0

    if not os.path.exists(folder_path):
        print(f"❌ Path not found: {folder_path}")
        sys.exit(1)

    for person in os.listdir(folder_path):
        person_path = os.path.join(folder_path, person)
        if not os.path.isdir(person_path):
            continue

        print(f"📂 Training from: {person_path}")
        label_map[label_id] = person

        for img_file in os.listdir(person_path):
            img_path = os.path.join(person_path, img_file)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if img is not None:
                faces.append(img)
                labels.append(label_id)
            else:
                print(f"⚠️ Unable to load: {img_path}")

        label_id += 1

    if not faces:
        print("❌ No valid images to train.")
        sys.exit(1)

    recognizer.train(faces, np.array(labels))
    return label_map


def mark_attendance(record_name, subject, branch, section):
    roll, name = record_name.split('_', 1)
    date_today = datetime.now().strftime('%Y-%m-%d')
    time_now = datetime.now().strftime('%H:%M:%S')

    folder = f"attendance/{date_today}"
    os.makedirs(folder, exist_ok=True)
    file_path = os.path.join(folder, f"{subject}_{branch}_{section}.xlsx")

    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Roll No", "Name", "Branch", "Section", "Date", "Time", "Status"])
        wb.save(file_path)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(values_only=True):
        if row[0] == roll and row[4] == date_today:
            print(f"🟡 Already marked present for {roll}")
            return

    sheet.append([roll, name, branch, section, date_today, time_now, "Present"])
    wb.save(file_path)
    print(f"✅ Attendance marked for {roll}")


def recognize_faces(label_map, subject, branch, section):
    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        if not ret:
            print("⚠️ Failed to capture from webcam.")
            continue

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            roi = gray[y:y + h, x:x + w]
            try:
                roi_resized = cv2.resize(roi, (200, 200))
                label, confidence = recognizer.predict(roi_resized)

                if confidence < 60:
                    person = label_map[label]
                    print(f"[✅ Recognized] {person} | Confidence: {confidence:.2f}")
                    mark_attendance(person, subject, branch, section)
                    cv2.putText(frame, f"{person}", (x, y - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    cv2.imshow("Attendance", frame)
                    cv2.waitKey(1000)
                    cap.release()
                    cv2.destroyAllWindows()
                    return
                else:
                    print(f"[❌ Not recognized] Confidence: {confidence:.2f}")
                    cv2.putText(frame, "Unknown", (x, y - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
            except Exception as e:
                print(f"❌ Error recognizing face: {e}")

        cv2.imshow("Attendance", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


# ✅ Main: accepts command-line arguments
if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("❌ Usage: python face_train.py <name_roll> <branch> <section>")
        sys.exit(1)

    name_roll = sys.argv[1]
    branch = sys.argv[2]
    section = sys.argv[3]

    # Create folder path and save images
    save_path = os.path.join("register_candidates", f"{branch}_{section}", name_roll)
    os.makedirs(save_path, exist_ok=True)

    cap = cv2.VideoCapture(0)
    count = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            continue

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face = gray[y:y+h, x:x+w]
            face = cv2.resize(face, (200, 200))
            file_path = os.path.join(save_path, f"{count}.png")
            cv2.imwrite(file_path, face)
            count += 1
            cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)

        cv2.imshow("Registering Face - Press 'q' to stop", frame)

        if cv2.waitKey(1) & 0xFF == ord('q') or count >= 20:
            break

    cap.release()
    cv2.destroyAllWindows()

    # Save registration to Excel
    os.makedirs("registered_data", exist_ok=True)
    # Create folder for this specific branch-section
    excel_dir = os.path.join("registered_data", f"{branch}_{section}")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, "students.xlsx")

# Create new Excel file if it doesn't exist
    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Roll No", "Name", "Branch", "Section", "Date", "Time"])
        wb.save(excel_path)

# Load and append new entry
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

# Check if already exists
    for row in sheet.iter_rows(values_only=True):
        if row[0] == roll and row[1] == name:
            print("🟡 Already registered.")
            break
    else:
        now = datetime.now()
        sheet.append([roll, name, branch, section, now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S')])
        wb.save(excel_path)
        print(f"✅ Registered in {excel_path}")


    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Roll No", "Name", "Class", "Section"])
        wb.save(excel_path)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    roll, name = name_roll.split('_', 1)
    ws.append([roll, name, branch, section])
    wb.save(excel_path)

    print("✅ Face registration complete.")

